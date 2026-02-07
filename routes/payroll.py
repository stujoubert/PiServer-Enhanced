#!/usr/bin/env python3
import os
import sqlite3
from dataclasses import dataclass, field
from datetime import datetime, date, timedelta, time
from typing import Dict, List, Optional

from flask import Blueprint, render_template, request, send_file
from dateutil import parser as dtparser

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from attendance.calc import calculate_daily_attendance
from services.user_helpers import list_users
from services.schedule_templates import get_user_schedule

from authz import login_required, role_required

# --------------------------------------------------
# Config
# --------------------------------------------------
DB_PATH = os.getenv("ATT_DB", "/var/lib/attendance/attendance.db")
bp = Blueprint("payroll", __name__, url_prefix="/payroll")

# --------------------------------------------------
# DB helpers
# --------------------------------------------------
def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

# --------------------------------------------------
# Time helpers
# --------------------------------------------------
def parse_ts(ts: str) -> Optional[datetime]:
    if not ts:
        return None
    try:
        dt = dtparser.parse(ts)
        if dt.tzinfo:
            dt = dt.astimezone().replace(tzinfo=None)
        return dt
    except Exception:
        return None

def fmt_hhmm(dt: Optional[datetime]) -> str:
    return dt.strftime("%H:%M") if dt else ""

def dec_hours_to_hhmm(hours: float) -> str:
    total_minutes = int(round(max(hours, 0) * 60))
    return f"{total_minutes // 60:02d}:{total_minutes % 60:02d}"

def daterange(d0: date, d1: date) -> List[date]:
    cur = d0
    out = []
    while cur <= d1:
        out.append(cur)
        cur += timedelta(days=1)
    return out

# --------------------------------------------------
# Schedule helpers (NO daily_hours, NO lunch)
# --------------------------------------------------
def parse_hhmm(t) -> datetime:
    """
    Accepts:
      - datetime.time
      - 'HH:MM'
      - 'HH:MM:SS'
    """
    if isinstance(t, time):
        return datetime.combine(date.today(), t)

    if isinstance(t, str):
        try:
            if len(t.split(":")) == 2:
                return datetime.strptime(t, "%H:%M")
            return datetime.strptime(t, "%H:%M:%S")
        except Exception:
            raise ValueError(f"Invalid time format: {t}")

    raise TypeError(f"Unsupported time type: {type(t)}")



def scheduled_seconds(schedule) -> Optional[int]:
    """
    Returns scheduled duration in seconds based ONLY on start/end.
    Handles overnight shifts.
    """
    if not schedule:
        return None

    start = schedule.get("start_time")
    end = schedule.get("end_time")

    if not start or not end:
        return None

    s = parse_hhmm(start)
    e = parse_hhmm(end)

    delta = (e - s).total_seconds()
    if delta < 0:
        delta += 24 * 3600  # overnight

    return int(delta)

# --------------------------------------------------
# Week helpers
# --------------------------------------------------
WEEK_TYPES = {
    "mon_sat": {"start_weekday": 0, "length_days": 6},
    "sat_fri": {"start_weekday": 5, "length_days": 7},
    "sun_sat": {"start_weekday": 6, "length_days": 7},
}

def week_start_for(day: date, week_type: str) -> date:
    cfg = WEEK_TYPES.get(week_type, WEEK_TYPES["mon_sat"])
    return day - timedelta(days=(day.weekday() - cfg["start_weekday"]) % 7)

def week_end_for(ws: date, week_type: str) -> date:
    cfg = WEEK_TYPES.get(week_type, WEEK_TYPES["mon_sat"])
    return ws + timedelta(days=cfg["length_days"] - 1)

def build_week_list(week_type: str, anchor: date, back=12, fwd=2):
    base = week_start_for(anchor, week_type)
    out = []
    for i in range(-back, fwd + 1):
        ws = base + timedelta(days=i * 7)
        out.append((ws, week_end_for(ws, week_type)))
    return out

# --------------------------------------------------
# Payroll computation
# --------------------------------------------------
@dataclass
class EmpRec:
    employee_id: str
    name: str
    days: Dict[str, dict] = field(default_factory=dict)
    total_regular: float = 0.0
    total_ot: float = 0.0
    total_all: float = 0.0
    # is_active is injected later in payroll_page/export

def _emp_sort_key(eid: str):
    return int(eid) if eid.isdigit() else eid

def compute_payroll(rows: List[sqlite3.Row], week_dates: List[date]):
    events_by_emp: Dict[str, Dict[date, List[dict]]] = {}

    # -----------------------------
    # Normalize events
    # -----------------------------
    for r in rows:
        emp_id = str(r["employee_id"] or "").strip()
        ts = parse_ts(r["timestamp"])
        if not emp_id or not ts:
            continue

        events_by_emp.setdefault(emp_id, {})
        events_by_emp[emp_id].setdefault(ts.date(), []).append({
            "employee_id": emp_id,
            "event_time": ts,
            "name": r["name"] or "",
        })

    results: List[EmpRec] = []

    # -----------------------------
    # Payroll per employee
    # -----------------------------
    for emp_id in sorted(events_by_emp.keys(), key=_emp_sort_key):
        daymap = events_by_emp[emp_id]

        emp_name = next(
            (e.get("name") for evs in daymap.values() for e in evs if e.get("name")),
            emp_id
        )

        emp = EmpRec(employee_id=emp_id, name=emp_name)

        reg_total = 0.0
        ot_total = 0.0

        for d in week_dates:
            evs = daymap.get(d, [])
            att = calculate_daily_attendance(evs)

            if not att:
                emp.days[d.isoformat()] = {
                    "in": "",
                    "out": "",
                    "hours": "00:00",
                    "ot": "00:00",
                    "hours_dec": 0.0,
                    "ot_dec": 0.0,
                    "flags": [],
                }
                continue

            worked_sec = att["worked_seconds"]
            flags = list(att.get("flags", []))

            # -----------------------------
            # Schedule resolution (FIXED)
            # -----------------------------
            try:
                schedule = get_user_schedule(emp_id, d.weekday())
            except Exception:
                schedule = None

            if schedule is None:
                # Truly no schedule assigned
                regular_sec = worked_sec
                ot_sec = 0
                flags.append("no_schedule")
            else:
                sched_sec = scheduled_seconds(schedule)

                if sched_sec is None:
                    # Schedule exists but malformed
                    regular_sec = worked_sec
                    ot_sec = 0
                    flags.append("invalid_schedule")
                else:
                    regular_sec = min(worked_sec, sched_sec)
                    ot_sec = max(worked_sec - sched_sec, 0)
                    if ot_sec > 0:
                        flags.append("overtime")

            rhrs = regular_sec / 3600
            ohrs = ot_sec / 3600

            reg_total += rhrs
            ot_total += ohrs

            emp.days[d.isoformat()] = {
                "in": fmt_hhmm(att["in"]),
                "out": fmt_hhmm(att["out"]),
                "hours": dec_hours_to_hhmm(rhrs),
                "ot": dec_hours_to_hhmm(ohrs),
                "hours_dec": round(rhrs, 2),
                "ot_dec": round(ohrs, 2),
                "flags": flags,
            }

        emp.total_regular = round(reg_total, 2)
        emp.total_ot = round(ot_total, 2)
        emp.total_all = round(reg_total + ot_total, 2)

        results.append(emp)

    return results


# --------------------------------------------------
# ROUTES
# --------------------------------------------------
@bp.route("/", methods=["GET"])
@login_required
@role_required("viewer", "manager", "admin")
def payroll_page():
    week_type = request.args.get("week_type", "mon_sat")
    user = request.args.get("user") or ""

    today = datetime.now().date()
    week_list = build_week_list(week_type, today)

    week_param = request.args.get("week")
    week_start = (
        dtparser.parse(week_param).date()
        if week_param
        else week_start_for(today, week_type)
    )

    week_end = week_end_for(week_start, week_type)
    week_dates = daterange(week_start, week_end)

    q_start = datetime.combine(week_start - timedelta(days=1), time.min)
    q_end = datetime.combine(week_end + timedelta(days=1), time.max)

    conn = get_conn()
    cur = conn.cursor()

    # -------------------------------------------------
    # 1. Fetch EVENTS ONLY (no logic change)
    # -------------------------------------------------
    sql = """
        SELECT e.employee_id, u.name, e.timestamp
        FROM events e
        JOIN users u ON u.employee_id = e.employee_id
        WHERE datetime(substr(e.timestamp,1,19))
              BETWEEN datetime(?) AND datetime(?)
    """
    params = [
        q_start.strftime("%Y-%m-%d %H:%M:%S"),
        q_end.strftime("%Y-%m-%d %H:%M:%S"),
    ]

    if user:
        sql += " AND e.employee_id = ?"
        params.append(user)

    rows = cur.execute(sql, params).fetchall()

    # -------------------------------------------------
    # 2. Build ACTIVE MAP (single query, safe)
    # -------------------------------------------------
    active_map = {
        str(r["employee_id"]): int(r["is_active"] or 0)
        for r in cur.execute(
            "SELECT employee_id, is_active FROM users"
        ).fetchall()
        if r["employee_id"] is not None
    }

    conn.close()

    # -------------------------------------------------
    # 3. Compute payroll (NOW NORMALIZED)
    # -------------------------------------------------
    payroll_data = compute_payroll(rows, week_dates)

    # -------------------------------------------------
    # 4. Inject is_active into payroll objects
    # -------------------------------------------------
    for emp in payroll_data:
        emp.is_active = active_map.get(str(emp.employee_id), 1)

    # -------------------------------------------------
    # 5. Render
    # -------------------------------------------------
    return render_template(
        "payroll.html",
        payroll_data=payroll_data,
        week_type=week_type,
        week_list=week_list,
        selected_week=week_start,
        week_start=week_start,
        week_end=week_end,
        week_dates=week_dates,
        users=list_users(),
        selected_user=user,
    )

@bp.route("/export", methods=["GET"])
def payroll_export():
    from flask import send_file, g, request
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from datetime import datetime, timedelta, time
    from dateutil import parser as dtparser

    week_type = request.args.get("week_type", "mon_sat")
    user = (request.args.get("user") or "").strip()

    today = datetime.now().date()
    week_param = request.args.get("week")
    week_start = dtparser.parse(week_param).date() if week_param else week_start_for(today, week_type)
    week_end = week_end_for(week_start, week_type)
    week_dates = daterange(week_start, week_end)

    q_start = datetime.combine(week_start - timedelta(days=1), time.min)
    q_end = datetime.combine(week_end + timedelta(days=1), time.max)

    conn = get_conn()
    cur = conn.cursor()

    # --------------------------------------------------
    # 1) Pull ALL users (source of truth for is_active)
    # --------------------------------------------------
    users_sql = """
        SELECT employee_id, name, COALESCE(is_active, 1) AS is_active
        FROM users
    """
    users_params = []
    if user:
        users_sql += " WHERE employee_id = ?"
        users_params.append(user)

    users_sql += " ORDER BY CAST(employee_id AS INTEGER) ASC"
    user_rows = cur.execute(users_sql, users_params).fetchall()

    # --------------------------------------------------
    # 2) Pull events in range (as before)
    # --------------------------------------------------
    events_sql = """
        SELECT
            e.employee_id,
            u.name,
            e.timestamp
        FROM events e
        JOIN users u ON u.employee_id = e.employee_id
        WHERE datetime(substr(e.timestamp,1,19))
              BETWEEN datetime(?) AND datetime(?)
    """

    events_params = [
        q_start.strftime("%Y-%m-%d %H:%M:%S"),
        q_end.strftime("%Y-%m-%d %H:%M:%S"),
    ]

    if user:
        events_sql += " AND employee_id = ?"
        events_params.append(user)

    event_rows = cur.execute(events_sql, events_params).fetchall()
    conn.close()

    # --------------------------------------------------
    # 3) Compute payroll ONLY for people with events
    # --------------------------------------------------
    computed = compute_payroll(event_rows, week_dates)
    computed_map = {str(emp.employee_id): emp for emp in computed}

    # --------------------------------------------------
    # 4) Build final ordered list: ALL users
    # --------------------------------------------------
    final_list = []
    for ur in user_rows:
        emp_id = str(ur["employee_id"])
        emp_name = ur["name"] or emp_id
        is_active = int(ur["is_active"] or 0)

        emp = computed_map.get(emp_id)
        if not emp:
            # Minimal blank object with same attributes used by the export
            class _BlankEmp:
                pass

            emp = _BlankEmp()
            emp.employee_id = emp_id
            emp.name = emp_name
            emp.days = {d.isoformat(): {
                "in": "",
                "out": "",
                "hours": "00:00",
                "ot": "00:00",
                "hours_dec": 0.0,
                "ot_dec": 0.0,
                "flags": [],
            } for d in week_dates}
            emp.total_regular = 0.0
            emp.total_ot = 0.0
            emp.total_all = 0.0

        # Always trust DB for active flag
        emp.is_active = is_active
        # Always trust DB for name (optional)
        if not getattr(emp, "name", ""):
            emp.name = emp_name

        final_list.append(emp)

    # --------------------------------------------------
    # 5) Excel generation (with freeze panes + header colors)
    # --------------------------------------------------
    wb = Workbook()
    ws = wb.active

    raw_title = g.T.get("payroll_title", "Payroll")
    ws.title = raw_title.replace("/", "-").replace("\\", "-")[:31]

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill("solid", fgColor="E9ECEF")   # light gray
    inactive_fill = PatternFill("solid", fgColor="FDECEA") # light red

    # Freeze first 2 rows and first 2 columns
    ws.freeze_panes = "C3"

    # Header Row 1
    ws.append([g.T.get("employee_id", "Employee ID"),
               g.T.get("name", "Employee Name")])

    # Style A1/B1
    for c in (1, 2):
        ws.cell(row=1, column=c).font = bold
        ws.cell(row=1, column=c).alignment = center
        ws.cell(row=1, column=c).fill = header_fill

    col = 3
    for d in week_dates:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 4)
        cell = ws.cell(row=1, column=col, value=d.strftime("%a %d/%m"))
        cell.font = bold
        cell.alignment = center
        cell.fill = header_fill
        # Ensure merged header region also gets fill
        for cc in range(col, col + 5):
            ws.cell(row=1, column=cc).fill = header_fill
        col += 5

    # Header Row 2
    ws.append(["", ""])
    # Style A2/B2
    for c in (1, 2):
        ws.cell(row=2, column=c).font = bold
        ws.cell(row=2, column=c).alignment = center
        ws.cell(row=2, column=c).fill = header_fill

    col = 3
    for _ in week_dates:
        labels = [
            g.T.get("in", "IN"),
            g.T.get("out", "OUT"),
            g.T.get("hours", "Hrs"),
            g.T.get("overtime", "OT"),
            "Flags",
        ]
        for i, label in enumerate(labels):
            c = ws.cell(row=2, column=col + i, value=label)
            c.font = bold
            c.alignment = center
            c.fill = header_fill
        col += 5

    # Data rows
    row_idx = 3
    for emp in final_list:
        ws.cell(row=row_idx, column=1, value=emp.employee_id)
        ws.cell(row=row_idx, column=2, value=emp.name)

        col = 3
        for d in week_dates:
            rec = emp.days.get(d.isoformat()) if getattr(emp, "days", None) else None
            ws.cell(row=row_idx, column=col,     value=rec["in"] if rec else "")
            ws.cell(row=row_idx, column=col + 1, value=rec["out"] if rec else "")
            ws.cell(row=row_idx, column=col + 2, value=timedelta(hours=(rec["hours_dec"] if rec else 0.0)))
            ws.cell(row=row_idx, column=col + 3, value=timedelta(hours=(rec["ot_dec"] if rec else 0.0)))
            ws.cell(row=row_idx, column=col + 4, value=", ".join(rec["flags"]) if rec and rec.get("flags") else "")
            col += 5

        ws.cell(row=row_idx, column=col,     value=timedelta(hours=float(getattr(emp, "total_regular", 0.0))))
        ws.cell(row=row_idx, column=col + 1, value=timedelta(hours=float(getattr(emp, "total_ot", 0.0))))
        ws.cell(row=row_idx, column=col + 2, value=timedelta(hours=float(getattr(emp, "total_all", 0.0))))

        # Highlight ONLY inactive users
        if int(getattr(emp, "is_active", 1)) == 0:
            last_col = col + 2
            for cc in range(1, last_col + 1):
                ws.cell(row=row_idx, column=cc).fill = inactive_fill

        row_idx += 1

    # Format timedelta cells as [h]:mm
    for col_cells in ws.iter_cols():
        for cell in col_cells:
            if isinstance(cell.value, timedelta):
                cell.number_format = "[h]:mm"

    # Optional: column widths for readability
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 28

    fname = f"weekly_payroll_{week_start}_to_{week_end}.xlsx"
    path = f"/tmp/{fname}"
    wb.save(path)

    return send_file(path, as_attachment=True, download_name=fname)
