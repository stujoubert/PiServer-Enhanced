from datetime import timedelta
from dateutil import parser as dtparser
from openpyxl import Workbook
from flask import g
from db import get_conn


def export_fifo_excel(output_path, start_date, end_date, week_dates=None):
    start_dt = dtparser.parse(start_date).date()
    end_dt = dtparser.parse(end_date).date()

    if week_dates is None:
        week_dates = []
        d = start_dt
        while d <= end_dt:
            week_dates.append(d)
            d += timedelta(days=1)

    conn = get_conn()
    cur = conn.cursor()
    rows = cur.execute(
        """
        SELECT
            e.employee_id,
            u.name AS name,
            DATE(e.timestamp) AS day,
            MIN(e.timestamp) AS first_in,
            MAX(e.timestamp) AS last_out
        FROM events e
        JOIN users u ON u.id = e.employee_id
        WHERE DATE(e.timestamp) BETWEEN DATE(?) AND DATE(?)
        GROUP BY e.employee_id, u.name, day
        ORDER BY CAST(e.employee_id AS INTEGER), day
        """,
        (start_dt.isoformat(), end_dt.isoformat()),
    ).fetchall()

    conn.close()

    data = {}
    for r in rows:
        emp = r["employee_id"]
        data.setdefault(emp, {"name": r["name"], "days": {}})
        data[emp]["days"][r["day"]] = {
            "in": dtparser.parse(r["first_in"]).strftime("%H:%M") if r["first_in"] else "",
            "out": dtparser.parse(r["last_out"]).strftime("%H:%M") if r["last_out"] else "",
        }

    T = getattr(g, "T", {})
    wb = Workbook()
    ws = wb.active
    ws.title = "FIFO Weekly"

    ws.append([T.get("employee_id", "Employee ID"), T.get("name", "Name")])

    col = 3
    for d in week_dates:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        ws.cell(row=1, column=col, value=d.strftime("%a %d/%m"))
        col += 2

    ws.append(["", ""])
    col = 3
    for _ in week_dates:
        ws.cell(row=2, column=col, value=T.get("first_in", "First IN"))
        ws.cell(row=2, column=col + 1, value=T.get("last_out", "Last OUT"))
        col += 2

    row_idx = 3
    for emp in sorted(data, key=lambda x: int(x) if str(x).isdigit() else x):
        ws.cell(row=row_idx, column=1, value=emp)
        ws.cell(row=row_idx, column=2, value=data[emp]["name"])
        col = 3
        for d in week_dates:
            rec = data[emp]["days"].get(d.isoformat())
            ws.cell(row=row_idx, column=col, value=rec["in"] if rec else "")
            ws.cell(row=row_idx, column=col + 1, value=rec["out"] if rec else "")
            col += 2
        row_idx += 1

    wb.save(output_path)
