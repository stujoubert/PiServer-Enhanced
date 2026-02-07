#!/usr/bin/env python3

import os
import sqlite3
import json
import subprocess
from datetime import datetime, timedelta
from dateutil import parser as dtparser
import urllib3
import uuid

urllib3.disable_warnings()

from flask import g

# IMPORTANT:
# Use the application's DB path consistently (ATT_DB) via db.py
from db import get_conn as _get_conn

# --------------------------------------------------
# Config
# --------------------------------------------------
DEVICE_TZ = os.getenv("HIK_TZ_OFFSET", "-06:00")


# --------------------------------------------------
# DB helper (single source of truth)
# --------------------------------------------------
def get_conn():
    return _get_conn()


# --------------------------------------------------
# Timestamp normalization
# --------------------------------------------------
def normalize_ts(ts: str) -> str:
    dt = dtparser.parse(ts)
    if dt.tzinfo:
        dt = dt.astimezone().replace(tzinfo=None)
    return dt.strftime("%Y-%m-%d %H:%M:%S")


# --------------------------------------------------
# EVENT FETCH
# --------------------------------------------------
def fetch_from_device(ip, username, password, start=None, end=None, device_id=None):
    if device_id is not None and (not start or not end):
        conn = get_conn()
        row = conn.execute(
            "SELECT last_fetch_at FROM devices WHERE id = ?",
            (device_id,)
        ).fetchone()
        conn.close()

        if row and row["last_fetch_at"]:
            start_dt = datetime.fromisoformat(row["last_fetch_at"]) - timedelta(minutes=2)
        else:
            start_dt = datetime.now() - timedelta(days=1)

        end_dt = datetime.now() + timedelta(minutes=1)
        start = start_dt.strftime("%Y-%m-%dT%H:%M:%S") + DEVICE_TZ
        end   = end_dt.strftime("%Y-%m-%dT%H:%M:%S") + DEVICE_TZ

    if not start or not end:
        raise ValueError("fetch_from_device requires start and end")

    url = f"http://{ip}/ISAPI/AccessControl/AcsEvent?format=json"

    all_events = []
    position = 0
    page_size = 50
    search_id = str(uuid.uuid4())

    while True:
        payload = {
            "AcsEventCond": {
                "searchID": search_id,
                "searchResultPosition": position,
                "maxResults": page_size,
                "major": 5,
                "minor": 75,
                "startTime": start,
                "endTime": end
            }
        }

        cmd = [
            "curl",
            "--digest",
            "-u", f"{username}:{password}",
            "-H", "Content-Type: application/json",
            "-d", json.dumps(payload),
            url
        ]

        result = subprocess.run(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True
        )

        if not result.stdout.strip():
            break

        try:
            data = json.loads(result.stdout)
        except Exception:
            break

        acs = data.get("AcsEvent", {})
        info = acs.get("InfoList", [])
        status = acs.get("responseStatusStrg")

        if not info:
            break

        for e in info:
            all_events.append({
                "employee_id": e.get("employeeNoString"),
                "name": e.get("name"),
                "timestamp": e.get("time"),
                "picture_url": e.get("pictureURL"),
            })

        position += len(info)
        if status != "MORE":
            break

    if device_id is not None and all_events:
        inserted = store_events(device_id, all_events)
        conn = get_conn()
        conn.execute(
            "UPDATE devices SET last_fetch_at = ?, last_fetch_count = ? WHERE id = ?",
            (max(e["timestamp"] for e in all_events), inserted, device_id)
        )
        conn.commit()
        conn.close()

    return all_events


# --------------------------------------------------
# STORE EVENTS
# --------------------------------------------------
def store_events(device_id, events):
    if not events:
        return 0

    conn = get_conn()
    cur = conn.cursor()
    inserted = 0

    for e in events:
        if not e.get("employee_id") or not e.get("timestamp"):
            continue
        try:
            norm_ts = normalize_ts(e["timestamp"])
            cur.execute(
                """
                INSERT OR IGNORE INTO events
                (device_id, employee_id, name, timestamp, picture_url)
                VALUES (?, ?, ?, ?, ?)
                """,
                (device_id, e["employee_id"], e.get("name"), norm_ts, e.get("picture_url"))
            )
            if cur.rowcount:
                inserted += 1
        except Exception:
            continue

    conn.commit()
    conn.close()
    return inserted


# --------------------------------------------------
# USERS FROM EVENTS
# --------------------------------------------------
def sync_users_from_events():
    conn = get_conn()
    cur = conn.cursor()

    rows = cur.execute(
        """
        SELECT DISTINCT
            e.employee_id,
            MAX(e.name) AS name
        FROM events e
        LEFT JOIN users u ON u.employee_id = e.employee_id
        WHERE e.employee_id IS NOT NULL
          AND u.employee_id IS NULL
        GROUP BY e.employee_id
        ORDER BY CAST(e.employee_id AS INTEGER)
        """
    ).fetchall()

    inserted = 0

    for r in rows:
        employee_id = r["employee_id"]
        name = r["name"] or f"Employee {employee_id}"

        try:
            cur.execute(
                "INSERT INTO users (employee_id, name) VALUES (?, ?)",
                (employee_id, name)
            )
            inserted += 1
        except Exception as e:
            print("User insert failed:", employee_id, e)

    conn.commit()
    conn.close()
    return inserted


# --------------------------------------------------
# USERS FROM DEVICE (existing behavior)
# --------------------------------------------------
def sync_missing_users_from_device(device_ip, user, password):
    import requests
    from requests.auth import HTTPDigestAuth
    from requests.exceptions import ConnectTimeout, ReadTimeout, ConnectionError

    conn = get_conn()
    c = conn.cursor()

    local_users = {
        row[0] for row in c.execute("SELECT employee_id FROM users").fetchall()
    }

    imported = 0
    skipped = 0
    position = 0
    page_size = 50

    while True:
        payload = {
            "UserInfoSearchCond": {
                "searchID": "1",
                "searchResultPosition": position,
                "maxResults": page_size,
            }
        }

        try:
            r = requests.post(
                f"http://{device_ip}/ISAPI/AccessControl/UserInfo/Search?format=json",
                json=payload,
                auth=HTTPDigestAuth(user, password),
                timeout=5,
                verify=False,
            )
        except (ConnectTimeout, ReadTimeout, ConnectionError):
            break

        if r.status_code != 200:
            break

        try:
            data = r.json()
        except Exception:
            break

        users = data.get("UserInfoSearch", {}).get("UserInfo", [])
        if not users:
            break

        for u in users:
            employee_id = u.get("employeeNo")
            name = u.get("name", "")
            if not employee_id or employee_id in local_users:
                skipped += 1
                continue

            c.execute(
                "INSERT INTO users (employee_id, name) VALUES (?, ?)",
                (employee_id, name)
            )
            local_users.add(employee_id)
            imported += 1

        if data["UserInfoSearch"].get("responseStatusStrg") != "MORE":
            break

        position += page_size

    conn.commit()
    conn.close()
    return {"imported": imported, "skipped": skipped}


# --------------------------------------------------
# FIFO EXPORT (restored; openpyxl imported lazily)
# --------------------------------------------------
def export_fifo_excel(output_path, start_date, end_date, week_dates=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    start_dt = dtparser.parse(start_date).date()
    end_dt = dtparser.parse(end_date).date()

    if week_dates is None:
        week_dates = []
        d = start_dt
        while d <= end_dt:
            week_dates.append(d)
            d += timedelta(days=1)

    conn = get_conn()
    cur = conn.cursor()
    rows = cur.execute(
        """
        SELECT employee_id, name, DATE(timestamp) AS day,
               MIN(timestamp) AS first_in,
               MAX(timestamp) AS last_out
        FROM events
        WHERE DATE(timestamp) BETWEEN DATE(?) AND DATE(?)
        GROUP BY employee_id, day
        ORDER BY CAST(employee_id AS INTEGER), day
        """,
        (start_dt.isoformat(), end_dt.isoformat()),
    ).fetchall()
    conn.close()

    data = {}
    for r in rows:
        emp = r["employee_id"]
        data.setdefault(emp, {"name": r["name"], "days": {}})
        data[emp]["days"][r["day"]] = {
            "in": dtparser.parse(r["first_in"]).strftime("%H:%M") if r["first_in"] else "",
            "out": dtparser.parse(r["last_out"]).strftime("%H:%M") if r["last_out"] else "",
        }

    T = getattr(g, "T", {}) or {}
    hdr_emp = T.get("employee_id", "Employee ID")
    hdr_name = T.get("name", "Name")
    hdr_in = T.get("first_in", "First IN")
    hdr_out = T.get("last_out", "Last OUT")

    wb = Workbook()
    ws = wb.active
    ws.title = "FIFO Weekly"
    ws.append([hdr_emp, hdr_name])

    col = 3
    for d in week_dates:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        ws.cell(row=1, column=col, value=d.strftime("%a %d/%m"))
        col += 2

    ws.append(["", ""])
    col = 3
    for _ in week_dates:
        ws.cell(row=2, column=col, value=hdr_in)
        ws.cell(row=2, column=col + 1, value=hdr_out)
        col += 2

    row_idx = 3
    for emp in sorted(data, key=lambda x: int(x) if str(x).isdigit() else x):
        ws.cell(row=row_idx, column=1, value=emp)
        ws.cell(row=row_idx, column=2, value=data[emp]["name"])
        col = 3
        for d in week_dates:
            rec = data[emp]["days"].get(d.isoformat())
            ws.cell(row=row_idx, column=col, value=rec["in"] if rec else "")
            ws.cell(row=row_idx, column=col + 1, value=rec["out"] if rec else "")
            col += 2
        row_idx += 1

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 30
    wb.save(output_path)


# --------------------------------------------
# FDLib helpers (unchanged behavior)
# --------------------------------------------
def fetch_fdlib_face(employee_id, device_ip, user, password):
    import requests
    from requests.auth import HTTPDigestAuth
    from requests.exceptions import ConnectTimeout, ReadTimeout, ConnectionError

    payload = {
        "searchResultPosition": 0,
        "maxResults": 1,
        "faceLibType": "blackFD",
        "FDID": "1",
        "FPID": str(employee_id),
    }

    try:
        r = requests.post(
            f"http://{device_ip}/ISAPI/Intelligent/FDLib/FDSearch?format=json",
            json=payload,
            auth=HTTPDigestAuth(user, password),
            timeout=5,
            verify=False,
        )
    except (ConnectTimeout, ReadTimeout, ConnectionError):
        return None

    if r.status_code != 200:
        return None

    try:
        data = r.json()
    except Exception:
        return None

    matches = data.get("MatchList") or []
    if not matches:
        return None

    return matches[0].get("faceURL")


def cache_fdlib_face(employee_id, face_url, user, password):
    import requests
    from requests.auth import HTTPDigestAuth

    FACE_DIR = "/var/lib/attendance/faces"
    os.makedirs(FACE_DIR, exist_ok=True)

    try:
        r = requests.get(
            face_url,
            auth=HTTPDigestAuth(user, password),
            timeout=5,
            verify=False,
        )
    except Exception:
        return None

    if r.status_code != 200 or not r.content:
        return None

    filename = f"{employee_id}.jpg"
    with open(os.path.join(FACE_DIR, filename), "wb") as f:
        f.write(r.content)

    return f"/users/faces/{filename}"


def persist_fdlib_face(employee_id, local_path):
    conn = get_conn()
    c = conn.cursor()
    c.execute(
        "INSERT OR REPLACE INTO user_faces (employee_id, picture_url, source_event_id) VALUES (?, ?, NULL)",
        (employee_id, local_path)
    )
    conn.commit()
    conn.close()


def bulk_import_fdlib_faces(device_ip, user, password):
    conn = get_conn()
    c = conn.cursor()
    users = c.execute("SELECT employee_id FROM users ORDER BY CAST(employee_id AS INTEGER)").fetchall()

    imported = 0
    skipped = 0

    for (employee_id,) in users:
        face_url = fetch_fdlib_face(employee_id, device_ip, user, password)
        if not face_url:
            skipped += 1
            continue

        local_path = cache_fdlib_face(employee_id, face_url, user, password)
        if not local_path:
            skipped += 1
            continue

        persist_fdlib_face(employee_id, local_path)
        imported += 1

    return {"imported": imported, "skipped": skipped}
